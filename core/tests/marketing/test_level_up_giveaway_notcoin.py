import os
from decimal import Decimal
from io import StringIO
from uuid import uuid4

import openpyxl
from django.core.files.temp import NamedTemporaryFile
from django.core.management import call_command
from django.test import TestCase

from exchange.accounts.models import Notification, User
from exchange.base.models import NOT_COIN
from exchange.wallet.models import Transaction, Wallet


class TestAzadUniGiveawayTests(TestCase):
    WEB_ENGAGE_IDS = (
        '9a12a59c-2d5e-4e87-aed3-ad429cec5274',
        'd3d5bead-2ffd-4f8d-b55f-3d5bee91c74a',
        '45f90342-2138-451d-bea4-2560fe15843f',
        '383e65b7-afc4-4da1-921c-6a707e924cac',
        '74bc3635-3fe0-49a5-9f3b-f198277708e6',
    )
    COMMAND = 'deposit_level_up_giveaways'
    NOT_FOUND_CUID = str(uuid4())
    SRC_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayLevelUp1403Src']
    DST_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayLevelUp1403Dst']
    BALANCES = (100, 200, 300, 400, 500)

    def setUp(self) -> None:
        self.excel_file = None

    def tearDown(self):
        if self.excel_file and os.path.exists(self.excel_file.name):
            os.remove(self.excel_file.name)

    def create_source_user_and_wallet(self, balance=1000):
        self.source_user = User.objects.create_user(username='financial@nobitex.ir')
        self.source_wallet = Wallet.objects.create(user=self.source_user, currency=NOT_COIN)
        self.source_wallet.balance = balance
        self.source_wallet.save(update_fields=['balance'])

    def create_user_and_wallets(self, count=5):
        self.users = []
        self.wallets = []

        for i in range(count):
            user = User.objects.create_user(
                username=f'test_user{i + 1}@nobitex.ir',
                webengage_cuid=self.WEB_ENGAGE_IDS[i],
            )
            wallet = Wallet.objects.create(user=user, currency=NOT_COIN)
            wallet.balance = self.BALANCES[i % len(self.BALANCES)]
            wallet.save(update_fields=['balance'])

            self.users.append(user)
            self.wallets.append(wallet)

    def create_test_excel_file(self, count=5, include_processed=False, include_not_found=False, include_over_50=False):
        rows = [
            [
                100 if include_over_50 else 50,
                self.NOT_FOUND_CUID if include_not_found else self.WEB_ENGAGE_IDS[0],
                'Processed' if include_processed else '',
            ],
            [50, self.WEB_ENGAGE_IDS[1], ''],
            [50, self.WEB_ENGAGE_IDS[2], ''],
            [50, self.WEB_ENGAGE_IDS[3], ''],
            [50, self.WEB_ENGAGE_IDS[4], ''],
        ]

        self.excel_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['amount', 'user id', 'status'])
        for i in range(count):
            sheet.append(rows[i])
        workbook.save(self.excel_file.name)
        workbook.close()

    def read_the_excel(self):
        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        return sheet.iter_rows(values_only=False, max_col=3)

    def test_bad_file(self):
        bad_file_path = '/tmp/bad_excel_file.xlsx'
        out = StringIO()

        call_command(self.COMMAND, bad_file_path, stdout=out)
        output = out.getvalue().strip()

        assert output == f'No such excel file: {bad_file_path}'

    def test_source_user_not_found(self):
        self.create_test_excel_file(count=0)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'The source user not found!'
        assert not Transaction.objects.exists()

    def test_empty_excel_file(self):
        self.create_test_excel_file(count=0)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'All records processed!\nprocessed 0 rows and 0 are already processed.'
        assert not Transaction.objects.exists()

    def test_over_50_notcoin(self):
        self.create_test_excel_file(count=1, include_over_50=True)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == (
            f'An error occurred for user_id={self.WEB_ENGAGE_IDS[0]}: Amount must be 50, but got 100\n'
            'All records processed!\n'
            'processed 0 rows and 0 are already processed.'
        )
        assert not Transaction.objects.exists()

    def test_including_processed(self):
        self.create_test_excel_file(count=1, include_processed=True)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == (
            'Skipping already processed record for user_id="9a12a59c-2d5e-4e87-aed3-ad429cec5274"'
            '\n'
            'All records processed!\nprocessed 0 rows and 1 are already processed.'
        )
        assert not Transaction.objects.exists()

    def test_including_not_found_user(self):
        self.create_test_excel_file(count=2, include_not_found=True)
        self.create_source_user_and_wallet()
        self.create_user_and_wallets(1)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == ('All records processed!\nprocessed 0 rows and 0 are already processed.')
        assert not Transaction.objects.exists()

        rows = self.read_the_excel()
        _ = next(rows)  # header
        _, cuid, status = next(rows)

        assert cuid.value == self.NOT_FOUND_CUID
        assert status.value == 'UserNotFound'

    def test_insufficient_balance_source(self):
        self.create_test_excel_file(count=2)
        self.create_source_user_and_wallet(balance=0)
        self.create_user_and_wallets(2)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == ('Insufficient the source wallet balance')
        assert not Transaction.objects.exists()

    def test_success(self):
        self.create_test_excel_file(count=2)
        self.create_source_user_and_wallet(balance=1000)
        self.create_user_and_wallets(2)
        out = StringIO()
        old_wallet_balance_user_1 = Decimal(self.wallets[0].balance)
        old_wallet_balance_user_2 = Decimal(self.wallets[1].balance)

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert (
            'Processed transaction for user=test_user1@nobitex.ir, '
            'webengage_id=9a12a59c-2d5e-4e87-aed3-ad429cec5274, '
            'amount=50NOT'
        ) in output

        assert (
            'Processed transaction for user=test_user2@nobitex.ir, '
            'webengage_id=d3d5bead-2ffd-4f8d-b55f-3d5bee91c74a, '
            'amount=50NOT'
        ) in output

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.withdraw,
            ).count()
            == 2
        )

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.deposit,
            ).count()
            == 2
        )

        assert (
            Notification.objects.filter(
                user__in=self.users,
                message__contains='جایزه احراز هویت شما در نوبیتکس',
            ).count()
            == 2
        )

        user_1_wallet = Wallet.objects.get(user=self.users[0], currency=NOT_COIN)
        assert user_1_wallet.balance == old_wallet_balance_user_1 + 50

        user_2_wallet = Wallet.objects.get(user=self.users[1], currency=NOT_COIN)
        assert user_2_wallet.balance == old_wallet_balance_user_2 + 50

        rows = self.read_the_excel()
        for row in rows:
            amount, user_id, status = row
            if amount.value == 'amount':
                continue
            assert status.value == 'Processed'
