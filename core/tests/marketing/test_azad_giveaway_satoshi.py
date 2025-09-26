import os
from decimal import Decimal
from io import StringIO
from uuid import uuid4

import openpyxl
from django.core.files.temp import NamedTemporaryFile
from django.core.management import call_command
from django.test import TestCase

from exchange.accounts.models import Notification, User
from exchange.base.models import Currencies
from exchange.wallet.models import Transaction, Wallet


class TestAzadUniGiveawayTests(TestCase):
    WEB_ENGAGE_IDS = (
        '9a12a59c-2d5e-4e87-aed3-ad429cec5274',
        'd3d5bead-2ffd-4f8d-b55f-3d5bee91c74a',
        '45f90342-2138-451d-bea4-2560fe15843f',
        '383e65b7-afc4-4da1-921c-6a707e924cac',
        '74bc3635-3fe0-49a5-9f3b-f198277708e6',
    )
    COMMAND = 'deposit_azad_giveaways'
    NOT_FOUND_CUID = str(uuid4())
    SRC_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayAzadUni1403Src']
    DST_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayAzadUni1403Dst']
    BALANCES = (0.01, 0.02, 0.03, 0.04, 0.05)
    REWARDS = (100, 200, 300, 400, 500)
    SATOSHI_IN_BTC = Decimal(100_000_000)
    BTC_PLACE = Decimal('.000001')

    def setUp(self) -> None:
        self.excel_file = None

    def tearDown(self):
        if self.excel_file and os.path.exists(self.excel_file.name):
            os.remove(self.excel_file.name)

    def create_source_user_and_wallet(self, balance=1):
        self.source_user = User.objects.create_user(username='financial@nobitex.ir')
        self.source_wallet = Wallet.objects.create(user=self.source_user, currency=Currencies.btc)
        self.source_wallet.balance = balance
        self.source_wallet.save(update_fields=['balance'])

    def create_user_and_wallets(self, count=5):
        self.users = []
        self.wallets = []

        for i in range(count):
            user = User.objects.create_user(
                username=f'test_user{i + 1}@nobitex.ir',
                webengage_cuid=self.WEB_ENGAGE_IDS[i],
            )
            wallet = Wallet.objects.create(user=user, currency=Currencies.btc)
            wallet.balance = self.BALANCES[i % len(self.BALANCES)]
            wallet.save(update_fields=['balance'])

            self.users.append(user)
            self.wallets.append(wallet)

    def create_test_excel_file(self, count=5, include_processed=False, include_not_found=False):
        rows = [
            [
                self.REWARDS[0],
                self.NOT_FOUND_CUID if include_not_found else self.WEB_ENGAGE_IDS[0],
                'Processed' if include_processed else '',
            ],
            [self.REWARDS[1], self.WEB_ENGAGE_IDS[1], ''],
            [self.REWARDS[2], self.WEB_ENGAGE_IDS[2], ''],
            [self.REWARDS[3], self.WEB_ENGAGE_IDS[3], ''],
            [self.REWARDS[4], self.WEB_ENGAGE_IDS[4], ''],
        ]

        self.excel_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['amount', 'user id', 'status'])
        for i in range(count):
            sheet.append(rows[i])
        workbook.save(self.excel_file.name)
        workbook.close()

    def read_the_excel(self):
        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        return sheet.iter_rows(values_only=False, max_col=3)

    def test_bad_file(self):
        bad_file_path = '/tmp/bad_excel_file.xlsx'
        out = StringIO()

        call_command(self.COMMAND, bad_file_path, stdout=out)
        output = out.getvalue().strip()

        assert output == f'No such excel file: {bad_file_path}'

    def test_source_user_not_found(self):
        self.create_test_excel_file(count=0)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'The source user not found!'
        assert not Transaction.objects.exists()

    def test_empty_excel_file(self):
        self.create_test_excel_file(count=0)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'All records processed!\nprocessed 0 rows and 0 are already processed.'
        assert not Transaction.objects.exists()

    def test_including_processed(self):
        self.create_test_excel_file(count=1, include_processed=True)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == (
            'Skipping already processed record for user_id="9a12a59c-2d5e-4e87-aed3-ad429cec5274"'
            '\n'
            'All records processed!\nprocessed 0 rows and 1 are already processed.'
        )
        assert not Transaction.objects.exists()

    def test_including_not_found_user(self):
        self.create_test_excel_file(count=2, include_not_found=True)
        self.create_source_user_and_wallet()
        self.create_user_and_wallets(2)
        out = StringIO()
        init_wallet_balance_source = Decimal(self.source_wallet.balance).quantize(self.BTC_PLACE)

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert (
            f'Processed transaction for user={self.users[1].username}, '
            f'webengage_id={self.WEB_ENGAGE_IDS[1]}, '
            f'amount=0.000002BTC'
        ) in output

        assert ('All records processed!' '\n' 'processed 1 rows and 0 are already processed.') in output

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.withdraw,
            ).count()
            == 1
        )

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.deposit,
            ).count()
            == 1
        )

        self.source_wallet.refresh_from_db()  # 0.999998
        reward_user = Decimal(self.REWARDS[1] / self.SATOSHI_IN_BTC).quantize(self.BTC_PLACE)  # 0.000002
        assert self.source_wallet.balance.quantize(self.BTC_PLACE) == (
            init_wallet_balance_source - reward_user
        ).quantize(self.BTC_PLACE)
        assert self.source_wallet.balance.quantize(self.BTC_PLACE) == Decimal('0.999998')

        rows = self.read_the_excel()
        _ = next(rows)  # header
        _, cuid, status = next(rows)
        assert cuid.value == self.NOT_FOUND_CUID
        assert status.value == 'UserNotFound'

        _, cuid, status = next(rows)
        assert cuid.value == self.WEB_ENGAGE_IDS[1]
        assert status.value == 'Processed'

    def test_insufficient_balance_source(self):
        self.create_test_excel_file(count=2)
        self.create_source_user_and_wallet(balance=0)
        self.create_user_and_wallets(2)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == ('Insufficient the source wallet balance')
        assert not Transaction.objects.exists()

    def test_success(self):
        self.create_test_excel_file(count=2)
        self.create_source_user_and_wallet(balance=1)
        self.create_user_and_wallets(2)
        out = StringIO()
        init_wallet_balance_user_1 = Decimal(self.wallets[0].balance).quantize(self.BTC_PLACE)
        init_wallet_balance_user_2 = Decimal(self.wallets[1].balance).quantize(self.BTC_PLACE)
        init_wallet_balance_source = Decimal(self.source_wallet.balance).quantize(self.BTC_PLACE)

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert (
            'Processed transaction for user=test_user1@nobitex.ir, '
            'webengage_id=9a12a59c-2d5e-4e87-aed3-ad429cec5274, '
            'amount=0.000001BTC'
        ) in output

        assert (
            'Processed transaction for user=test_user2@nobitex.ir, '
            'webengage_id=d3d5bead-2ffd-4f8d-b55f-3d5bee91c74a, '
            'amount=0.000002BTC'
        ) in output

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.withdraw,
            ).count()
            == 2
        )

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.deposit,
            ).count()
            == 2
        )

        assert (
            Notification.objects.filter(
                user__in=self.users,
                message__contains='مشارکت شما در مسابقه برگزار شده در دانشگاه آزاد',
            ).count()
            == 2
        )

        user_1_wallet = Wallet.objects.get(user=self.users[0], currency=Currencies.btc)
        # convert reward of user1 to satoshi
        reward_user1 = Decimal(self.REWARDS[0] / self.SATOSHI_IN_BTC).quantize(self.BTC_PLACE)
        assert user_1_wallet.balance == init_wallet_balance_user_1 + reward_user1
        assert user_1_wallet.balance == Decimal('0.010001')

        user_2_wallet = Wallet.objects.get(user=self.users[1], currency=Currencies.btc)
        reward_user2 = Decimal(self.REWARDS[1] / self.SATOSHI_IN_BTC).quantize(self.BTC_PLACE)
        assert user_2_wallet.balance == init_wallet_balance_user_2 + reward_user2
        assert user_2_wallet.balance == Decimal('0.020002')

        self.source_wallet.refresh_from_db()
        assert self.source_wallet.balance.quantize(self.BTC_PLACE) == (
            init_wallet_balance_source - (reward_user1 + reward_user2)
        ).quantize(self.BTC_PLACE)
        assert self.source_wallet.balance.quantize(self.BTC_PLACE) == Decimal('0.999997')

        rows = self.read_the_excel()
        for row in rows:
            amount, user_id, status = row
            if amount.value == 'amount':
                continue
            assert status.value == 'Processed'
