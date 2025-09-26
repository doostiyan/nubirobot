import os
from decimal import Decimal
from io import StringIO
from uuid import uuid4

import openpyxl
from django.core.files.temp import NamedTemporaryFile
from django.core.management import call_command
from django.test import TestCase

from exchange.accounts.models import Notification, User
from exchange.base.models import Currencies
from exchange.wallet.models import Transaction, Wallet


class TestHamsterPrelaunchGiveawayTests(TestCase):
    WEB_ENGAGE_IDS = (
        '9a12a59c-2d5e-4e87-aed3-ad429cec5274',
        'd3d5bead-2ffd-4f8d-b55f-3d5bee91c74a',
        '45f90342-2138-451d-bea4-2560fe15843f',
        '383e65b7-afc4-4da1-921c-6a707e924cac',
        '74bc3635-3fe0-49a5-9f3b-f198277708e6',
    )
    COMMAND = 'deposit_hamster_ton_giveaways'
    NOT_FOUND_CUID = str(uuid4())
    SRC_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayHamsterPreLaunchSrc']
    DST_TRANSACTION_REF_MODULE = Transaction.REF_MODULES['CampaignGiveawayHamsterPreLaunchDst']
    BALANCES = (100, 200, 300, 400, 500)

    def setUp(self) -> None:
        self.excel_file = None

    def tearDown(self):
        if self.excel_file and os.path.exists(self.excel_file.name):
            os.remove(self.excel_file.name)

    def create_source_user_and_wallet(self, balance=1000):
        self.source_user = User.objects.create_user(username='financial@nobitex.ir')
        self.source_wallet = Wallet.objects.create(user=self.source_user, currency=Currencies.ton)
        self.source_wallet.balance = balance
        self.source_wallet.save(update_fields=['balance'])

    def create_user_and_wallets(self, start_row=0, end_row=0):
        self.users = []
        self.wallets = []

        for i in range(start_row, end_row):
            user = User.objects.create_user(
                username=f'test_user{i + 1}@nobitex.ir',
                webengage_cuid=self.WEB_ENGAGE_IDS[i],
            )
            wallet = Wallet.objects.create(user=user, currency=Currencies.ton)
            wallet.balance = self.BALANCES[i % len(self.BALANCES)]
            wallet.save(update_fields=['balance'])

            self.users.append(user)
            self.wallets.append(wallet)

    def create_test_excel_file(self, include_processed=False, include_not_found=False, start_row=0, end_row=0):
        default_rows = [
            [self.WEB_ENGAGE_IDS[0], ''],
            [self.WEB_ENGAGE_IDS[1], ''],
            [self.WEB_ENGAGE_IDS[2], ''],
            [self.WEB_ENGAGE_IDS[3], ''],
            [self.WEB_ENGAGE_IDS[4], ''],
        ]
        adding_non_default_row = include_processed or include_not_found
        if adding_non_default_row:
            rows = [
                [
                    self.NOT_FOUND_CUID if include_not_found else self.WEB_ENGAGE_IDS[0],
                    'Processed' if include_processed else '',
                ],
                *default_rows[start_row : end_row - 1],
            ]
        else:
            rows = default_rows[start_row:end_row]

        self.excel_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        workbook.save(self.excel_file.name)
        workbook.close()

    def read_the_excel(self):
        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        return sheet.iter_rows(values_only=False, max_col=2)

    def test_bad_file(self):
        bad_file_path = '/tmp/bad_excel_file.xlsx'
        out = StringIO()

        call_command(self.COMMAND, bad_file_path, stdout=out)
        output = out.getvalue().strip()

        assert output == f'No such excel file: {bad_file_path}'

    def test_source_user_not_found(self):
        self.create_test_excel_file(start_row=0, end_row=0)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'The source user not found!'
        assert not Transaction.objects.exists()

    def test_empty_excel_file(self):
        self.create_test_excel_file(start_row=0, end_row=0)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()

        assert output == 'All records processed!\nprocessed 0 rows and 0 are already processed.'
        assert not Transaction.objects.exists()

    def test_including_processed(self):
        self.create_test_excel_file(include_processed=True, start_row=1, end_row=2)
        self.create_source_user_and_wallet()
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == (
            'Skipping already processed record for user_id="9a12a59c-2d5e-4e87-aed3-ad429cec5274"'
            '\n'
            'All records processed!\nprocessed 0 rows and 1 are already processed.'
        )
        assert not Transaction.objects.exists()

    def test_including_not_found_user(self):
        self.create_test_excel_file(include_not_found=True, start_row=0, end_row=1)
        self.create_source_user_and_wallet()
        self.create_user_and_wallets(start_row=0, end_row=1)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == ('All records processed!\nprocessed 0 rows and 0 are already processed.')
        assert not Transaction.objects.exists()

        rows = self.read_the_excel()
        cuid, status = next(rows)

        assert cuid.value == self.NOT_FOUND_CUID
        assert status.value == 'UserNotFound'

    def test_insufficient_balance_source(self):
        self.create_test_excel_file(start_row=0, end_row=2)
        self.create_source_user_and_wallet(balance=0)
        self.create_user_and_wallets(start_row=0, end_row=2)
        out = StringIO()

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert output == ('Insufficient the source wallet balance')
        assert not Transaction.objects.exists()

    def test_success(self):
        self.create_test_excel_file(start_row=2, end_row=4)
        self.create_source_user_and_wallet(balance=1000)
        self.create_user_and_wallets(start_row=2, end_row=4)
        out = StringIO()
        old_wallet_balance_user_1 = Decimal(self.wallets[0].balance)
        old_wallet_balance_user_2 = Decimal(self.wallets[1].balance)

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert (
            'Processed transaction for user=test_user3@nobitex.ir, '
            'webengage_id=45f90342-2138-451d-bea4-2560fe15843f, '
            'amount=1TON'
        ) in output

        assert (
            'Processed transaction for user=test_user4@nobitex.ir, '
            'webengage_id=383e65b7-afc4-4da1-921c-6a707e924cac, '
            'amount=1TON'
        ) in output

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.withdraw,
            ).count()
            == 2
        )

        assert (
            Transaction.objects.filter(
                tp=Transaction.TYPE.deposit,
            ).count()
            == 2
        )

        assert (
            Notification.objects.filter(
                user__in=self.users,
                message='کاربر گرامی، جایزه گوش‌به‌زنگ شما در نوبیتکس به میزان 1 تون‌کوین'
                ' (TON) '
                'به کیف پول نوبیتکس شما واریز شد.\n'
                'می‌توانید برای واریز آن به تون‌کیپر اقدام کنید.',
            ).count()
            == 2
        )

        user_1_wallet = Wallet.objects.get(user=self.users[0], currency=Currencies.ton)
        assert user_1_wallet.balance == old_wallet_balance_user_1 + 1

        user_2_wallet = Wallet.objects.get(user=self.users[1], currency=Currencies.ton)
        assert user_2_wallet.balance == old_wallet_balance_user_2 + 1

        rows = self.read_the_excel()
        for row in rows:
            user_id, status = row
            assert status.value == 'Processed'

    def test_duplicate_ref_id(self):
        Transaction.objects.all().delete()
        self.create_test_excel_file(start_row=4, end_row=5)
        self.create_source_user_and_wallet(balance=1000)
        self.create_user_and_wallets(start_row=4, end_row=5)
        out = StringIO()
        old_wallet_balance_user_1 = Decimal(self.wallets[0].balance)

        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert (
            'Processed transaction for user=test_user5@nobitex.ir, '
            'webengage_id=74bc3635-3fe0-49a5-9f3b-f198277708e6, '
            'amount=1TON'
        ) in output
        user_1_wallet = Wallet.objects.get(user=self.users[0], currency=Currencies.ton)
        assert user_1_wallet.balance == old_wallet_balance_user_1 + 1

        out = StringIO()
        self.create_test_excel_file(start_row=4, end_row=5)
        call_command(self.COMMAND, self.excel_file.name, stdout=out)
        output = out.getvalue().strip()
        assert (
            'An error occurred for user_id=74bc3635-3fe0-49a5-9f3b-f198277708e6: '
            'duplicate key value violates unique constraint '
            '"wallet_transaction_ref_module_ref_id_810dbbe3_uniq"\n'
            f'DETAIL:  Key (ref_module, ref_id)=(673775432, {self.users[0].id}) already exists.\n'
            'All records processed!\n'
            'processed 0 rows and 0 are already processed.'
        ) in output
        user_1_wallet.refresh_from_db()
        assert user_1_wallet.balance == old_wallet_balance_user_1 + 1
