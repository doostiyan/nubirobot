import os
from tempfile import NamedTemporaryFile
from unittest.mock import patch

import openpyxl
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.db import IntegrityError
from django.test import TestCase

from exchange.accounts.models import BankAccount, User
from exchange.features.models import QueueItem
from exchange.shetab.models import JibitAccount, JibitPaymentId


class AddJibitPaymentIdCommandTest(TestCase):
    COMMAND_NAME = 'make_ideposit_payids_vip_users'

    def create_test_excel_file(self, include_wrong_account=False, include_processed=False, empty_row=False):
        rows = [
            [
                'NJ1001',
                '1002003001',
                'IR760120000000007565000016',
                '7565000016',
                'Processed' if include_processed else '',
            ],
            ['NJ1002', '1002003002', 'IR760120000000007565000016', '7565000016', ''],
            [
                'NJ1003',
                '1002003003',
                'IR760120000000007565000016',
                'wrong_account' if include_wrong_account else '7565000016',
                '',
            ],
            ['NJ1004', '1002003004', 'IR760120000000007565000016', '7565000016', ''],
            ['NJ1005', '1002003005', 'IR760120000000007565000016', '7565000016', ''],
        ]
        if empty_row:
            rows.append([' ', '', '', '', ''])

        self.excel_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['reference_number', 'pay_id', 'iban', 'account', 'status'])
        for row in rows:
            sheet.append(row)
        workbook.save(self.excel_file.name)
        workbook.close()

    def setUp(self):
        self.excel_file = None
        self.user1, _ = User.objects.get_or_create(username='user_1@nobitex.ir')
        self.user2, _ = User.objects.get_or_create(username='user_2@nobitex.ir')
        self.user3, _ = User.objects.get_or_create(username='user_3@nobitex.ir')
        self.bank_account1 = BankAccount.objects.create(
            id=1001,
            user=self.user1,
            shaba_number='IR123456789012345678901234',
            confirmed=True,
        )
        self.bank_account2 = BankAccount.objects.create(
            id=1002,
            user=self.user2,
            shaba_number='IR234567890123456789012345',
            confirmed=True,
        )
        self.bank_account3 = BankAccount.objects.create(
            id=1003,
            user=self.user3,
            shaba_number='IR345678901234567890123456',
            confirmed=True,
        )

        self.jibit_account = JibitAccount.objects.create(
            bank=JibitAccount.BANK_CHOICES.BKMTIR,
            iban='IR760120000000007565000016',
            account_number='7565000016',
            owner_name='راهکار فناوری نویان',
            account_type=JibitAccount.ACCOUNT_TYPES.nobitex_jibit,
        )

        self.feature_key = QueueItem.FEATURES.nobitex_jibit_ideposit

    def tearDown(self):
        if self.excel_file and os.path.exists(self.excel_file.name):
            os.remove(self.excel_file.name)
        QueueItem.objects.all().delete()
        JibitPaymentId.objects.all().delete()

    def run_command(self):
        call_command(self.COMMAND_NAME, self.excel_file.name)

    def test_successful_creation_with_correct_queue_item(self):
        self.create_test_excel_file()
        self.jibit_account.delete()
        self.run_command()
        assert JibitPaymentId.objects.filter(payment_id='1002003001').exists()
        assert JibitPaymentId.objects.filter(payment_id='1002003002').exists()
        assert JibitPaymentId.objects.filter(payment_id='1002003003').exists()

        assert QueueItem.objects.filter(
            user=self.bank_account1.user,
            feature=self.feature_key,
            status=QueueItem.STATUS.done,
        ).exists()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for index, row in enumerate(rows):
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            if index > 2:  # 3 record with related bank account was successfully
                assert status_cell.value == 'BankAccountNotFound'

    def test_skip_processed_row(self):
        self.create_test_excel_file(include_processed=True)
        self.run_command()

        assert not JibitPaymentId.objects.filter(payment_id='1002003001').exists()
        assert not QueueItem.objects.filter(user=self.bank_account1.user, feature=self.feature_key).exists()

    def test_destination_account_mismatch(self):
        self.create_test_excel_file(include_wrong_account=True)
        self.run_command()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for index, row in enumerate(rows):
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            if index < 2:
                assert status_cell.value == 'Processed'
            elif index == 2:
                assert status_cell.value == 'DestinationAccountNotMatched'
            else:
                assert status_cell.value == 'BankAccountNotFound'

        assert QueueItem.objects.filter(feature=self.feature_key, status=QueueItem.STATUS.done).count() == 2
        assert not QueueItem.objects.filter(
            feature=self.feature_key,
            user=self.user3,
            status=QueueItem.STATUS.done,
        ).exists()
        assert JibitPaymentId.objects.all().count() == 2

    def test_bank_account_not_found(self):
        self.bank_account1.delete()
        self.bank_account2.delete()
        self.bank_account3.delete()
        self.create_test_excel_file()
        self.run_command()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for row in rows:
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            assert status_cell.value == 'BankAccountNotFound'

        assert not JibitPaymentId.objects.all().exists()
        assert not QueueItem.objects.filter(feature=self.feature_key).exists()

    def test_reference_number_invalid(self):
        rows = [
            ['NAXXX', '1002003001', 'IR760120000000007565000016', '7565000016', ''],
        ]

        self.excel_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['reference_number', 'pay_id', 'iban', 'account', 'status'])
        for row in rows:
            sheet.append(row)
        workbook.save(self.excel_file.name)
        workbook.close()

        self.run_command()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for row in rows:
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            assert status_cell.value == 'ReferenceNumberInvalid'

        assert not QueueItem.objects.filter(feature=self.feature_key).exists()
        assert not JibitPaymentId.objects.all().exists()

    def test_feature_activation_failure(self):
        self.create_test_excel_file()

        with patch(
            'exchange.features.models.QueueItem.enable_feature',
            side_effect=Exception('Feature activation error'),
        ):
            self.run_command()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for index, row in enumerate(rows):
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            if index < 3:
                assert status_cell.value == 'CantActivateFeatureFlag'
            else:
                assert status_cell.value == 'BankAccountNotFound'

        assert not JibitPaymentId.objects.all().exists()
        assert not QueueItem.objects.filter(feature=self.feature_key).exists()

    def test_integrity_error_handling(self):
        self.create_test_excel_file()

        with patch('exchange.shetab.models.JibitPaymentId.objects.get_or_create', side_effect=IntegrityError):
            self.run_command()

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for index, row in enumerate(rows):
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            if index < 3:
                assert status_cell.value == 'AlreadyExist'
            else:
                assert status_cell.value == 'BankAccountNotFound'

        assert JibitPaymentId.objects.count() == 0
        assert QueueItem.objects.filter(feature=self.feature_key).count() == 0

    def test_account_type_validation(self):
        jibit_account = JibitAccount(
            iban='IR123456789012345678901234',
            account_number='123456789',
            owner_name='test',
            account_type=10,
        )
        with self.assertRaises(ValidationError):
            jibit_account.full_clean()

        jibit_account.account_type = JibitAccount.ACCOUNT_TYPES.jibit
        jibit_account.full_clean()

    def test_empty_row(self):
        self.create_test_excel_file(empty_row=True)
        self.run_command()
        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False, min_row=2, max_col=5)
        for index, row in enumerate(rows):
            (
                reference_number_cell,
                pay_id_cell,
                destination_iban_cell,
                destination_account_cell,
                status_cell,
            ) = row
            if index <= 2:
                assert status_cell.value == 'Processed'
            elif index < 5:
                assert status_cell.value == 'BankAccountNotFound'
            elif index == 5:
                assert status_cell.value is None
